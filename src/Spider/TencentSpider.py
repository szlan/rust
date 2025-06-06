import requests
from jsonpath import jsonpath
from openpyxl import Workbook, load_workbook
import time
import random
import os
from bs4 import BeautifulSoup

def get_data(url, headers, payload):
    try:
        r = requests.post(url, headers=headers, json=payload)
        if r.status_code == 200:
            return r.json()
        else:
            print(f"请求失败，状态码: {r.status_code}")
            return None
    except Exception as e:
        print(f"请求发生异常: {e}")
        return None

def get_published_time(url, headers):
    try:
        r = requests.get(url, headers=headers)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')
            meta_tag = soup.find('meta', {'property': 'article:published_time'})
            if meta_tag:
                return meta_tag.get('content')
        return None
    except Exception as e:
        print(f"获取发布时间时发生异常: {e}")
        return None

def jxx(data, category, headers, counter):
    if not data:
        return counter
    title = jsonpath(data, '$..title')
    url2 = jsonpath(data, '$..url')
    tag_words = jsonpath(data, '$..chl_name')
    for titles, urls, tag_wordss in zip(title, url2, tag_words):
        if counter >= 5:  # 如果该分类已经爬取了5条新闻，停止处理
            break
        published_time = get_published_time(urls, headers)
        print(f'新闻标题：{titles}\n新闻链接：{urls}\n新闻媒体：{tag_wordss}\n新闻分类：{category}\n发布时间：{published_time}')
        print('==============' * 5)
        save_data(titles, urls, tag_wordss, category, published_time)
        counter += 1  # 每保存一条新闻，计数器加1
    return counter

def save_data(tit, lin, name, category, published_time):
    ws.append([tit, lin, name, category, published_time])
    wb.save(excel_file_path)

if __name__ == '__main__':
    # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file_path = os.path.join(script_dir, '腾讯新闻_全分类.xlsx')
    
    # 检查文件是否已存在
    if os.path.exists(excel_file_path):
        # 加载已存在的工作簿
        wb = load_workbook(excel_file_path)
        ws = wb.active
        print(f"已加载现有文件: {excel_file_path}")
    else:
        # 创建新工作簿并添加表头
        wb = Workbook()
        ws = wb.active
        ws.append(['新闻标题', '新闻链接', '新闻媒体', '新闻分类', '发布时间'])
        print(f"创建新文件: {excel_file_path}")
    
    url = 'https://i.news.qq.com/web_feed/getHotModuleList'
    headers = {
        'cookie': 'RK=gwvBNF6tO3; ptcz=730ae62e10f461f434b2938d6e74f25acb2859f6640602ee4b646bbdfbd8f675; pgv_pvid=1250429676; _qimei_uuid42=1890c0e2f25100a84caa1f52e4120aa3392af06f5f; pac_uid=0_SM2WDY82ADn3S; current-city-name=fuzhou; _qimei_fingerprint=0a18e34c9ba93dc545680d1dbbd15eac; _qimei_q36=; _qimei_h38=e147f4b94caa1f52e4120aa30200000381890c; suid=user_0_SM2WDY82ADn3S; lcad_o_minduid=SWtwqZV2_IvpuYTEOQGHVEynZuRUaDJl; lcad_appuser=60957B3A3CC1B5A5; lcad_Lturn=392; lcad_LKBturn=933; lcad_LPVLturn=654; lcad_LPLFturn=400; lcad_LPSJturn=594; lcad_LBSturn=629; lcad_LVINturn=485; lcad_LDERturn=289',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36 Edg/112.0.1722.58'
    }

    category_ids = {
        '要闻': 'news_news_top',
        '财经': 'news_news_finance',
        '科技': 'news_news_tech',
        '娱乐': 'news_news_ent',
        '体育': 'news_news_sports',
        '国际': 'news_news_world',
        '军事': 'news_news_mil',
        '汽车': 'news_news_auto',
        '教育': 'news_news_edu',
        '健康': 'news_news_antip',
        '游戏': 'news_news_game',
        '科学': 'news_news_kepu',
        '历史': 'news_news_history'
    }

    for category, channel_id in category_ids.items():
        print(f"开始爬取 {category} 分类的新闻...")
        counter = 0  # 初始化计数器
        page = 1
        while counter < 5:  # 当该分类的新闻数量小于5时，继续爬取
            payload = {
                'se_req': {'from': "pc"},
                'forward': "2",
                'qimei36': "0_SM2WDY82ADn3S",
                'device_id': "0_SM2WDY82ADn3S",
                'base_req': {'from': "pc"},
                'channel_id': channel_id,
                'flush_num': page,
                'item_count': 20,
            }
            json_data1 = get_data(url, headers, payload)
            time.sleep(random.randint(2, 6))
            counter = jxx(json_data1, category, headers, counter)  # 更新计数器
            page += 1  # 翻页
    
    print(f"数据已成功保存到: {excel_file_path}")